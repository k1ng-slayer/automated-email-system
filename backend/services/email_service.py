import smtplib
from contextlib import contextmanager
from dataclasses import dataclass, field
from email.message import EmailMessage
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


ALLOWED_EXTENSIONS = {"xlsx", "xlsm", "xlsb", "xltx", "xltm", "xls", "xlt", "xml"}
SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 587


def is_allowed_excel(filename: str) -> bool:
    if "." not in filename:
        return False
    return filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


@dataclass
class SendResult:
    total_rows: int = 0
    processed_rows: int = 0
    sent_count: int = 0
    skipped_count: int = 0
    failures: list[str] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "total_rows": self.total_rows,
            "processed_rows": self.processed_rows,
            "sent_count": self.sent_count,
            "skipped_count": self.skipped_count,
            "failures": self.failures,
        }


def _build_defaulter_message(
    rollno: str | int | None,
    fullname: str | None,
    sender: str,
    receiver: str,
    attendance: float,
    additional_message: str,
) -> EmailMessage:
    email = EmailMessage()
    email["From"] = sender
    email["To"] = receiver
    email["Subject"] = "Attendance for last month"
    body = (
        f"Roll No:{rollno} {fullname}, your attendance for last month is {attendance}%. "
        f"\n{additional_message}"
    )
    email.set_content(body)
    return email


def _build_customize_message(
    rollno: str | int | None,
    fullname: str | None,
    sender: str,
    receiver: str,
    subject: str,
    custom_message: str,
) -> EmailMessage:
    email = EmailMessage()
    email["From"] = sender
    email["To"] = receiver
    email["Subject"] = subject
    body = f"Greetings {fullname}!\n{custom_message}"
    email.set_content(body)
    return email


def _load_sheet(workbook_bytes: bytes, sheet_name: str) -> Worksheet:
    workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {available}"
        )
    return workbook[sheet_name]


@contextmanager
def _smtp_session(sender_email: str, sender_password: str):
    smtp = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)
    try:
        smtp.starttls()
        smtp.login(sender_email, sender_password)
        yield smtp
    finally:
        try:
            smtp.quit()
        except smtplib.SMTPException:
            pass


def send_emails(
    *,
    mode: str,
    workbook_bytes: bytes,
    sender_email: str,
    sender_password: str,
    message: str = "",
    threshold_attendance: float | None = None,
    subject: str = "",
    sheet_name: str = "Sheet1",
) -> SendResult:
    sheet = _load_sheet(workbook_bytes, sheet_name=sheet_name)
    result = SendResult()

    rows = sheet.max_row
    cols = sheet.max_column

    if rows < 2:
        return result
    if mode not in {"defaulter", "customize"}:
        raise ValueError("Invalid mode. Use 'defaulter' or 'customize'.")
    if cols < 4:
        raise ValueError(
            "Customize sheet must have at least 4 columns: Roll No, Name, Class, Email."
        )
    if mode == "defaulter":
        if cols < 5:
            raise ValueError(
                "Defaulter sheet must have 5 columns: Roll No, Name, Class, Email, Attendance."
            )
        if threshold_attendance is None:
            raise ValueError("Threshold attendance is required for defaulter mode.")

    result.total_rows = max(rows - 1, 0)

    with _smtp_session(sender_email, sender_password) as smtp:
        for row in range(2, rows + 1):
            result.processed_rows += 1
            rollno = sheet.cell(row=row, column=1).value
            fullname = sheet.cell(row=row, column=2).value
            receiver_raw = sheet.cell(row=row, column=4).value
            receiver = str(receiver_raw).strip() if receiver_raw is not None else ""

            if not receiver:
                result.skipped_count += 1
                result.failures.append(f"Row {row}: missing receiver email.")
                continue

            try:
                if mode == "defaulter":
                    attendance_raw = sheet.cell(row=row, column=5).value
                    if attendance_raw is None:
                        result.skipped_count += 1
                        result.failures.append(f"Row {row}: missing attendance.")
                        continue

                    try:
                        attendance = float(attendance_raw)
                    except (TypeError, ValueError):
                        result.skipped_count += 1
                        result.failures.append(
                            f"Row {row}: invalid attendance value '{attendance_raw}'."
                        )
                        continue

                    if attendance >= threshold_attendance:
                        result.skipped_count += 1
                        continue

                    outgoing = _build_defaulter_message(
                        rollno=rollno,
                        fullname=str(fullname) if fullname is not None else "",
                        sender=sender_email,
                        receiver=receiver,
                        attendance=attendance,
                        additional_message=message,
                    )
                else:
                    outgoing = _build_customize_message(
                        rollno=rollno,
                        fullname=str(fullname) if fullname is not None else "",
                        sender=sender_email,
                        receiver=receiver,
                        subject=subject,
                        custom_message=message,
                    )

                smtp.send_message(outgoing)
                result.sent_count += 1
            except smtplib.SMTPException:
                result.failures.append(f"Row {row}: SMTP send failed.")

    return result
